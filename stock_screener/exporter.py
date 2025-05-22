# stock_screener/exporter.py
'''
auto_adjust_columns: Adjust automatically the dimensions of the columns
apply_conditional_formatting: Draw cells either in red or green depending on the value
export_to_excel: Export datas into excel alike files and apply the previous functions
'''

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment

def auto_adjust_columns(ws):
    max_len=[]
    for col in ws.columns:
        """
        max_len = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2
        """
        for cell in col:
            # Appliquer l'alignement vertical centré et le retour à la ligne automatique
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            # Ajuster la hauteur de la ligne en fonction du contenu
            if cell.value:
                max_len.append(len(str(cell.value)))
                lines = str(cell.value).count('\n') + 1
                # Estimer la hauteur de la ligne (approximativement 50 points par ligne)
                estimated_height = lines * 30
                current_height = ws.row_dimensions[cell.row].height
                if current_height is None or estimated_height > current_height:
                    ws.row_dimensions[cell.row].height = estimated_height
        if (ws._WorkbookChild__title == 'Sommaire') or (col[0].column_letter == 'A'):
            ws.column_dimensions[col[0].column_letter].width = int(max(max_len)/1.5) + 2

'''
def apply_conditional_formatting(ws):
    red = PatternFill("solid", fgColor="FFC7CE")
    green = PatternFill("solid", fgColor="C6EFCE")
    # format >1 en rouge, ≤1 en vert
    ws.conditional_formatting.add(
        f"B2:{ws.max_column}{ws.max_row}",
        CellIsRule(operator="greaterThan", formula=["1"], fill=red)
    )
    ws.conditional_formatting.add(
        f"B2:{ws.max_column}{ws.max_row}",
        CellIsRule(operator="lessThanOrEqual", formula=["1"], fill=green)
    )
'''
def export_to_excel(summary: pd.DataFrame, details: dict, filename: str):
    """
    summary : DataFrame pour l'onglet 'Sommaire'
    details : {sheet_name: DataFrame_ratios}
    """
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Sommaire", index=False)
        for name, df in details.items():
            df.to_excel(writer, sheet_name=name[:31])
    wb = load_workbook(filename)
    for ws in wb.worksheets:
        auto_adjust_columns(ws)
        #apply_conditional_formatting(ws)
    wb.save(filename)
