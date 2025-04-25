import pandas as pd
import os
from stock_screener.exporter import export_to_excel

def test_exporter(tmp_path):
    summary = pd.DataFrame([{"Nom":"X","TAG":"X","ISIN":"Y","Prix":1.0}])
    details = {"X": pd.DataFrame({"r":[1,2]}, index=[2022,2023])}
    out = tmp_path/"out.xlsx"
    export_to_excel(summary, details, str(out))
    assert out.exists()
    import openpyxl
    wb = openpyxl.load_workbook(str(out))
    assert "Sommaire" in wb.sheetnames
    assert "X" in wb.sheetnames
