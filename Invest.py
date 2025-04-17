import os
import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
import matplotlib.pyplot as plt
import datetime as dt

#
def fetch_company_data(company_identifiers):
    """Récupère les données financières pour une société donnée."""
    base_url = "https://www.zonebourse.com/cours/action/"
    end_url = "valorisation/"
    url = os.path.join(base_url,company_identifiers,end_url)
#
    # Utilisation d'un header pour simuler un navigateur et éviter d'être bloqué
    headers = {
        "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) "
                       "Chrome/115.0 Safari/537.36")
    }
#
    response = requests.get(url, headers=headers)
    if response.status_code != 200:
        print(f"Erreur : Impossible d'accéder aux données pour {company_identifiers} (HTTP {response.status_code})")
        return None
#
    soup = BeautifulSoup(response.text, 'html.parser')
#######################################
    end_url2="finances-ratios/"
    url2 = os.path.join(base_url,company_identifiers,end_url2)
    response2 = requests.get(url2, headers=headers)
    soup2 = BeautifulSoup(response2.text, 'html.parser')
############################################################################
    TAG=soup.find_all('h2',attrs='m-0 badge txt-b5 txt-s1')[0].text.strip()
    ISIN=soup.find_all('h2',attrs='m-0 badge txt-b5 txt-s1')[1].text.strip()
    prix_devise=soup.find_all('td',attrs='txt-s7 txt-align-left is__realtime-last')[0].text.strip() #Avec devise
    prix=prix_devise.split()[0]
#
    tag_tbody=soup.find('tbody') #corps du tableau
    tag_tr=tag_tbody.find_all('tr') #l'ensemble des lignes du tableau
    indexe_ratios = [0,2,4,5,6,9,11,13,15,17,18,20,21,23] #Ratios intéressants
    val_par_ligne = len(tag_tr[indexe_ratios[0]].find_all('td')) #Nb element par ligne hors titre
    #ligne_capitalisation=tag_tr[indexe_ratios[0]].find_all('td')
#LISTE des NOMS
    liste_noms=tag_tbody.find_all("td",attrs={'class':'table-child--w200'})
    titre_capitalisation=liste_noms[indexe_ratios[0]].text.strip().split('\n')[0]
    #
    titre_valorisation=liste_noms[indexe_ratios[1]].text.strip().split('\n')[0]
    #
    titre_PER=liste_noms[indexe_ratios[2]].text.strip().split('\n')[0]
    titre_PBR=liste_noms[indexe_ratios[3]].text.strip().split('\n')[0]
    titre_PEG=liste_noms[indexe_ratios[4]].text.strip().split('\n')[0]
    #
    #
    titre_EV_EBITDA=liste_noms[indexe_ratios[5]].text.strip().split('\n')[0]
    #
    titre_EV_FCF=liste_noms[indexe_ratios[6]].text.strip().split('\n')[0]
    #
    titre_dividendeParAction=liste_noms[indexe_ratios[7]].text.strip().split('\n')[0]
    #
    titre_BNA=liste_noms[indexe_ratios[8]].text.strip().split('\n')[0]
    #
    titre_CA=liste_noms[indexe_ratios[9]].text.strip().split('\n')[0]
    titre_EBITDA=liste_noms[indexe_ratios[10]].text.strip().split('\n')[0]
    #
    titre_beneficeNet=liste_noms[indexe_ratios[11]].text.strip().split('\n')[0]
    titre_endettementNet=liste_noms[indexe_ratios[12]].text.strip().split('\n')[0]
    #
    titre_nbTitre=liste_noms[indexe_ratios[13]].text.strip().split('\n')[0]
#Liste des VALEURS
    dico_valeurs = {
        titre_capitalisation:[],
        titre_valorisation:[],
        titre_PER:[],
        titre_PBR:[],
        titre_PEG:[],
        titre_EV_EBITDA:[],
        titre_EV_FCF:[],
        titre_dividendeParAction:[],
        titre_BNA:[],
        titre_CA:[],
        titre_EBITDA:[],
        titre_beneficeNet:[],
        titre_endettementNet:[],
        titre_nbTitre:[],
    }
    i=0
    for name in dico_valeurs.keys():
        values=[]
        for j in range(1,val_par_ligne):
            try:
                test=float(tag_tr[indexe_ratios[i]].find_all('td')[j].text.strip().replace('\u202f','').replace('x','').replace(',','.'))
            except ValueError:
                #Mise à 0 pour ne pas avoir de problème de graphique, df etc... mais embrouille la lecture des ratios... trouver un autre moyen
                test=float(tag_tr[indexe_ratios[i]].find_all('td')[j].text.strip().replace('-','0'))
            finally:
                values.append(test)
        dico_valeurs[name]=values
        i=i+1
    i=0
    #Muliplication des lignes par 1000000
    #Methode alternative:   map(lambda x: x * 1E6, data['Capitalisation'])
    dico_valeurs['Capitalisation'] = [(1E6)*x for x in dico_valeurs['Capitalisation']]
    dico_valeurs['Valeur Entreprise'] = [(1E6)*x for x in dico_valeurs['Valeur Entreprise']]
    dico_valeurs["Chiffre d'affaires"] = [(1E6)*x for x in dico_valeurs["Chiffre d'affaires"]]
    dico_valeurs['EBITDA'] = [(1E6)*x for x in dico_valeurs['EBITDA']]
    dico_valeurs['Résultat net'] = [(1E6)*x for x in dico_valeurs['Résultat net']]
    dico_valeurs["Endettement Net"] = [(1E6)*x for x in dico_valeurs['Endettement Net']]
    dico_valeurs['Nbr de Titres (en Milliers)'] = [(1E6)*x for x in dico_valeurs['Nbr de Titres (en Milliers)']]
    #Modifier la clef sur le nombre de titre
    dico_valeurs['Nbr de Titres']=dico_valeurs.pop('Nbr de Titres (en Milliers)')
################################################################
        #Mise en df
        #df=pd.DataFrame.from_dict(dico_valeurs,orient='index')
        #Multiplication de 1000000 pour les lignes concernées
        #df.loc['Capitalisation',:] *= 1E6
        #df.loc['Valeur Entreprise',:] *= 1E6
        #df.loc["Chiffre d'affaires",:] *= 1E6
        #df.loc['EBITDA',:] *= 1E6
        #df.loc['Résultat net',:] *= 1E6
        #df.loc["Endettement Net",:] *= 1E6
        #df.loc['Nbr de Titres (en Milliers)',:] *= 1E6
###################################################################
    Date=[]
    Date_head=soup.find('thead')
    Date_section=Date_head.find_all('span')
    for k in range(len(Date_section)):
        date_scrap=int(Date_section[k].text.strip())
        Date.append(date_scrap)

    df=pd.DataFrame(dico_valeurs).T
    df.columns = Date
###################################################################
###################################################################
    #Ratios complémentaires issue de url2
    tag_tbody2=soup2.find_all('tbody') #corps du tableau
    tag_tr2=tag_tbody2[3].find_all('tr') #l'ensemble des lignes du tableau
    liste_noms2=tag_tbody2[3].find_all("p",attrs={'class':'c txt-inline table-child--hover-display m-0 ml-5 pl-0'})
    indexe_ratios2 = [0,2,10,19,26] #Ratios intéressants
    indexe_ratios3 = [1,3,12,23,31]
    ##############
    if liste_noms2[indexe_ratios2[2]].text.strip().split('\n')[0] != "Marge nette %":
        indexe_ratios2 = [0,2,9,18,25]
        indexe_ratios3 = [1,3,11,22,30] #Erreur de placement de ratios sur certaines entreprises
    else:
        pass
############
    val_par_ligne2 = len(tag_tr2[indexe_ratios3[0]].find_all('td',attrs={'table-child--right table-child--w80'})) #Nb element par ligne hors titre
    rentabiliteActifs=liste_noms2[indexe_ratios2[0]].text.strip().split('\n')[0]
    rentabiliteCapitauxPropres=liste_noms2[indexe_ratios2[1]].text.strip().split('\n')[0]
    margeNette=liste_noms2[indexe_ratios2[2]].text.strip().split('\n')[0]
    liquiditeCourtTerme=liste_noms2[indexe_ratios2[3]].text.strip().split('\n')[0]
    solvabiliteLongTerme=liste_noms2[indexe_ratios2[4]].text.strip().split('\n')[0]

    dico_calculs = {
        rentabiliteActifs:[],
        rentabiliteCapitauxPropres:[],
        #"titre_ROIC":[], FLEMMMMMMMMME
        #"titre_PCF":[], FLEEEEEEEEEEEEEEEME
        margeNette:[],
        liquiditeCourtTerme:[],
        solvabiliteLongTerme:[],
    }

    i=0
    for name in dico_calculs.keys():
        values=[]
        for j in range(6,val_par_ligne2+6):
            try:
                test=float(tag_tr2[indexe_ratios3[i]].find_all('td')[j].text.strip().replace('\u202f','').replace('x','').replace(',','.'))
            except ValueError:
                #Mise à 0 pour ne pas avoir de problème de graphique, df etc... mais embrouille la lecture des ratios... trouver un autre moyen
                test=float(tag_tr2[indexe_ratios3[i]].find_all('td')[j].text.strip().replace('-','0'))
            finally:
                values.append(test)
        dico_calculs[name]=values
        i=i+1
    i=0

    Date2=[]
    Date_head2=soup2.find('thead')
    Date_section2=Date_head2.find_all('span')
    for k in range(5,len(Date_section2)):
        date_scrap=int(Date_section2[k].text.strip())
        Date2.append(date_scrap)

    df2=pd.DataFrame(dico_calculs).T
    df2.columns = Date2

    df3=pd.concat([df,df2], ignore_index=False) #Combine les deux df

    return TAG, ISIN, prix, prix_devise, df3
#
""" def main():
    company_identifiers = ["SAFRAN-4696/valorisation/", "FERMENTALG-16118042/valorisation/", "MANITOU-GROUP-4773/valorisation/"]
    output_file = "donnees_financieres.txt"

    with open(output_file, "w", encoding="utf-8") as file:
        for company in company_identifiers:
            print(f"Traitement de {company}...")
            data = fetch_company_data(company)
            if data:
                file.write(f"{company}:\n")
                for key, value in data.items():
                    file.write(f"  {key}: {value}\n")
                file.write("\n")
            else:
                file.write(f"{company}: Erreur lors de la récupération des données.\n")

if __name__ == "__main__":
    main() """

# Fonction pour ajuster automatiquement la taille des cellules sur Excel
def ajuster_largeur_colonnes(ws):
    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max_length + 2

# Création des graphiques
def plot_ratios(df, company_name):
    os.makedirs("graphs", exist_ok=True)

    # Graphique 1 : Valorisation (PER & PEG)
    df.T[["PER", "PEG"]].plot(marker='o', figsize=(8, 5), title=f"Valorisation - {company_name}")
    plt.ylabel("-")
    plt.xlabel("Année")
    plt.grid(True)
    plt.legend()
    plt.tight_layout()
    plt.savefig(f"graphs/{company_name}_PER_PEG.png")
    plt.close()

    # Graphique 2 : Gestion des ressources (ROE, ROA, Marge nette)
    df.T[["Rentabilité des actifs (ROA)", "Rentabilité des actifs (ROA)", "Marge nette %"]].plot(marker='o', figsize=(8, 5), title=f"Gestion des ressources - {company_name}")
    plt.ylabel("-")
    plt.xlabel("Année")
    plt.grid(True)
    plt.legend()
    plt.tight_layout()
    plt.savefig(f"graphs/{company_name}_ROE_ROA_MargeNette.png")
    plt.close()

    # Graphique 3 : Maîtrise des coûts (Debt/Equity, Current ratio, EV/EBITDA)
    df.T[["Total des dettes/capitaux propres", "Current Ratio", "Valeur Entreprise / EBITDA"]].plot(marker='o', figsize=(8, 5), title=f"Maîtrise des coûts - {company_name}")
    plt.ylabel("-")
    plt.xlabel("Année")
    plt.grid(True)
    plt.legend()
    plt.tight_layout()
    plt.savefig(f"graphs/{company_name}_DebtEquity_CurrentRatio_EVEBITDA.png")
    plt.close()

def main():
    current_year=int(dt.date.today().strftime('%Y'))
    company_identifiers = [
        "SAFRAN-4696/",
        "FERMENTALG-16118042/",
        "MANITOU-GROUP-4773/",
        "NVIDIA-CORPORATION-57355629/",
        "TOTALENERGIES-SE-4717/",
        "BYD-COMPANY-LIMITED-5640763/",
        "ASML-HOLDING-N-V-12002973/",
        "NOVO-NORDISK-A-S-1412980/",
        "LVMH-4669/"
        #"BNP-PARIBAS-4618/",
        #"AIR-LIQUIDE-4605/",
        #"SAINT-GOBAIN-4697/",
        #"TSMC-TAIWAN-SEMICONDUCTOR-6492349/",
        #"RHEINMETALL-AG-436527/",
        #"SANOFI-4698/",
        #"MEDPACE-HOLDINGS-INC-30506552/"
    ]
    output_file = "donnees_financieres.xlsx"

    Sommaire = []

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for company in company_identifiers:
            nom_compagnie = company.split('-')[0].strip().upper()
            print(f"Traitement de {nom_compagnie}...")

            try:
                TAG, ISIN, prix, prix_devise, df3 = fetch_company_data(company)
            except Exception as e:
                print(f"Erreur lors du traitement de {nom_compagnie}: {e}")
                continue

#Création d'un dictionnaire pour le sommaire
            Sommaire.append({
                "Nom": nom_compagnie,
                "TAG": TAG,
                "ISIN": ISIN,
                "Prix": prix_devise,
                "Dividende par action" : df3.loc["Dividende / Action",current_year]
            })
#Analyse fondamentale de la compagnie par onglet
            df3.to_excel(writer, sheet_name=nom_compagnie[:31])  # Excel limite à 31 caractères
#Mise en place des graphes
            plot_ratios(df3, nom_compagnie)
#mise en df du dico Sommaire puis écriture dans l'onglet Sommaire
        resume_df = pd.DataFrame(Sommaire)
        resume_df.to_excel(writer, sheet_name="Sommaire", index=False)


# Ajustement des colonnes dans excel
    wb = load_workbook(output_file)
    for ws in wb.worksheets:
        ajuster_largeur_colonnes(ws)
    wb.save(output_file)

    print(f"Données exportées dans '{output_file}' avec succès.")

if __name__ == "__main__":
    main()